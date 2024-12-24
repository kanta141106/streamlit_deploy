import openpyxl
import json
import openai
import os

from dotenv import load_dotenv

load_dotenv()
from openpyxl import Workbook
import streamlit as st

import tempfile
from io import BytesIO

structure_map = {
    "機器番号": 2,   
    "機器名称": 3,    
    "形式": 4,
    "系統名": 5,
    "設置場所": 6,
    "台数":7,
    "風量":{
        "SA・SOA(m3/h)":8,
        "RA(m3/h)":9,
        "OA(m3/h)":10
        },
    "静圧":11,
    "シングルコイル":{
        "冷却能力":{
            "冷却能力(kw)":12,
            "入口空気":{
                "DB":13,
                "WB":14
                },
            "出口空気":{
                "DB":15,
                "WB":16
            },
            "冷水":{
                "入口":17,
                "出口":18,
                "流量":19
            },
        },
        "加熱能力":{
            "加熱能力(kw)":20,
            "入口空気":{
                "DB":21
            },
            "出口空気":{
                "DB":22
            },
            "温水":{
                "入口":23,
                "出口":24,
                "流量":25
            },
        },
        "コイル形式":26,
        "コイル列数":27
    },
    "加湿器":{
        "方式":28,
        "有効加湿量":29
    },
    "フィルター":{
        "プレフィルター":30,
        "メインフィルター":31
    },
    "防振（ファン部）":33,
    "電気特性":{
        "電源(φ-V)":34,
        "ファン(kW)":35
        },
    "インバータ":36,
    "寸法":{
        "W":37,
        "D":38,
        "H":39
        },
    "製品重量(kg)":40,
    "備考":41
}

def fill_data(sheet, structure, col):
    """
    structure が:
      - int: Excelの行番号 → 値を取得
      - dict: 再帰的に要素を辿る
    """
    if isinstance(structure, int):
        # 行番号を示す → セル値を取得
        row_number = structure
        return sheet.cell(row=row_number, column=col).value
    
    elif isinstance(structure, dict):
        # 辞書の場合はキーごとに再帰
        result = {}
        for key, sub_map in structure.items():
            result[key] = fill_data(sheet, sub_map, col)
        return result
    
    else:
        # 予想外の型ならそのまま返す等 (必要ならエラーにしても良い)
        return structure

def sheet1(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    
    sheet = wb["Table 1"]  # 読み込むシートを指定
    all_units_data = []
    # max_col = sheet.max_column

    # 列番号のリスト
    start_col = 12
    max_col = sheet.max_column

    # 列ループ
    for col in range(start_col, max_col + 1):
        # まず 2行目からユニット名を取得
        unit_name = sheet.cell(row=2, column=col).value
        if not unit_name:
            # ユニット名が空(None)ならスキップするなど
            continue
        
        # スキーマに従ってデータを埋め込む
        coil_data = fill_data(sheet, structure_map, col)
        
        # まとめて辞書を作る
        unit_dict = {
        }
        # "コイル情報" を unit_dict に統合する
        unit_dict.update(coil_data)
        
        # リストに追加
        all_units_data.append(unit_dict)

    # JSONファイルに書き出し
    with open("output.json", "w", encoding="utf-8") as f:
        json.dump(all_units_data, f, ensure_ascii=False, indent=2)

    print("Done.")

    # OpenAI APIキーを環境変数から取得
    openai.api_key = st.secrets["openai"]["api_key"]
    # client = openai.OpenAI()
    # 入力JSONファイルのパスを指定
    input_json_path = "output.json"

    # JSONファイルを読み込む
    with open(input_json_path, "r") as file:
        input_json = json.load(file)

    # プロンプトとパラメータの設定
    prompt = f"""
以下のJSONデータから必要な情報を抽出し、新しいJSON形式に変換してください。
入力されたjsonファイルには機器ごとに情報がまとまっています。以下のようなjson形式にしてください。
機器情報
    - 「機器名」には「機器名称」を使用してください。
    - 「メーカー」は「備考欄」を参照して、メーカー名が書いてあれば入力し、書いてなければ空白のまま飛ばしてください。
    - 「設備種類」は以下の「設備種類、セクション、サブセクションの決め方」をもとに入力してください。
    - 「セクション」は以下の「設備種類、セクション、サブセクションの決め方」をもとに入力してください。
    - 「サブセクション」は以下の「設備種類、セクション、サブセクションの決め方」をもとに入力してください。
    - 「型式」には「形式」を使用してください。
    - 「型番」には「型番」を使用し、英数字部分だけを抜き出してください
    - 「設置場所」には「設置場所」をもとに使用してください。「\n」は省略してください
    - 「屋内・屋外」は「設置場所」や「備考」を参考にして、屋内、屋外、不明のいずれかを設定してください。
    - 「数量」は「台数」の値を使用してください。
    - メモにはその他の項目にない情報（例えば、冷却能力、加熱能力、風量、消費電力）を記載してください。
###設備種類、セクション、サブセクションの決め方
1.まずそれぞれの機器情報をもとに以下のキーワードから今回最適な項目を選んでください
キーワード：
    • 冷温水発生機
    • 冷凍機
    • 外調機
    • 排気ファン
    • 給気ファン
    • 全熱交換器
2.選んだキーワードに対応する「設備種類」、「セクション」、「サブセクション」を値として入力してください。
- キーワード：冷温水発生機
    - 設備種類：空調設備
    - セクション：熱源機器
    - サブセクション：冷温水発生器
- キーワード：冷凍機
    - 設備種類：空調設備
    - セクション：熱源機器
    - サブセクション：冷凍機
- キーワード：外調機
    - 設備種類：空調設備
    - セクション：空調機
    - サブセクション：外調機
- キーワード：排気ファン
    - 設備種類：空調設備
    - セクション：：給排気設備
    - サブセクション：排気ファン
- キーワード：給気ファン
    - 設備種類：空調設備
    - セクション：給排気設備
    - サブセクション：給気ファン
- キーワード：全熱交換器
    - 設備種類：空調設備
    - セクション：給排気設備
    - サブセクション：全熱交換器

###jsonデータ
以下のJSONデータを基に処理してください: {input_json}
    """
    chat_completion = openai.ChatCompletion.create(
            messages=[{
                "role": "user",
                "content": prompt
            }],
            model="gpt-4o",
            temperature=0.1,
            response_format={"type": "json_object"}
        )
    a=json.loads(chat_completion.choices[0].message.content)

    wb = Workbook()
    ws = wb.active
    ws.title = "機器情報"

    # 列名を定義
    columns = ["機器名", "メーカー", "設備種類", "セクション", "サブセクション", "型式", "型番", "設置場所", "屋内・屋外", "数量", "メモ"]
    # 列名を1行目に設定
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # データを行ごとに記入
    current_row = 2  # データは2行目から記入
    for machine in a['機器情報']:
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=current_row, column=col_idx, value=machine.get(col_name, ""))
        current_row += 1

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

#シート2
structure_map = {
    "機器番号":2,
    "機器名称":3,
    "形式":7,
    "系統名":8,
    "設置場所":9,
    "番手":10,
    "台数":11,
    "冷房能力":{
        "顕熱能力":12,
        "全熱能力":13,
    },
    "暖房能力":14,
    "コイル仕様":15,
    "加湿":{
        "方式":16,
        "加湿量":17,
    },
    "送風機":{
        "参考風量":18,
        "機外静圧":19,
    },
    "冷水":{
        "入口温度":20,
        "出口温度":21,
        "流量":22,
    },
    "温水":{
        "入口温度":23,
        "出口温度":24,
        "流量":25
    },
    "消費電力":{
        "電源":26,
        "冷房":27,
        "暖房":28,
    },
    "フィルター":29,
    "防振":30,
    "連動":31,
    "騒音値":32,
    "寸法":{
        "W":33,
        "D":34,
        "H":35
    },
    "重量":36,
    "型番":37,
    "備考":38
}

def sheet2(path):
    # Excelファイルを読み込み
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb["Table 2"]  # 読み込むシートを指定
    def fill_data(sheet, structure, col):
        """
        structure が:
        - int: Excelの行番号 → 値を取得
        - dict: 再帰的に要素を辿る
        """
        if isinstance(structure, int):
            # 行番号を示す → セル値を取得
            row_number = structure
            return sheet.cell(row=row_number, column=col).value
        
        elif isinstance(structure, dict):
            # 辞書の場合はキーごとに再帰
            result = {}
            for key, sub_map in structure.items():
                result[key] = fill_data(sheet, sub_map, col)
            return result
        
        else:
            # 予想外の型ならそのまま返す等 (必要ならエラーにしても良い)
            return structure

    all_units_data = []
    # max_col = sheet.max_column

    # 列番号のリスト
    start_col = 5
    max_col = sheet.max_column

    # 列ループ
    for col in range(start_col, max_col + 1):
        # まず 2行目からユニット名を取得
        unit_name = sheet.cell(row=2, column=col).value
        if not unit_name:
            # ユニット名が空(None)ならスキップするなど
            continue
        
        # スキーマに従ってデータを埋め込む
        coil_data = fill_data(sheet, structure_map, col)
        
        # まとめて辞書を作る
        unit_dict = {
        }
        # "コイル情報" を unit_dict に統合する
        unit_dict.update(coil_data)
        
        # リストに追加
        all_units_data.append(unit_dict)

    # JSONファイルに書き出し
    with open("output_1.json", "w", encoding="utf-8") as f:
        json.dump(all_units_data, f, ensure_ascii=False, indent=2)

    print("Done.")

    # OpenAI APIキーを環境変数から取得
    openai.api_key = st.secrets["openai"]["api_key"]
    # client = openai.OpenAI()
    # 入力JSONファイルのパスを指定
    input_json_path = "output_1.json"

    # JSONファイルを読み込む
    with open(input_json_path, "r") as file:
        input_json = json.load(file)

    # プロンプトとパラメータの設定
    prompt = f"""
以下のJSONデータから必要な情報を抽出し、新しいJSON形式に変換してください。
入力されたjsonファイルには機器ごとに情報がまとまっています。以下のようなjson形式にしてください。
機器情報
    - 「機器名」には「機器名称」を使用してください。
    - 「メーカー」は「備考欄」を参照して、メーカー名が書いてあれば入力し、書いてなければ空白のまま飛ばしてください。
    - 「設備種類」は以下の「設備種類、セクション、サブセクションの決め方」をもとに入力してください。
    - 「セクション」は以下の「設備種類、セクション、サブセクションの決め方」をもとに入力してください。
    - 「サブセクション」は以下の「設備種類、セクション、サブセクションの決め方」をもとに入力してください。
    - 「型式」には「形式」を使用してください。
    - 「型番」には「型番」を使用し、英数字部分だけを抜き出してください
    - 「設置場所」には「設置場所」をもとに使用してください。「\n」は省略してください
    - 「屋内・屋外」は「設置場所」や「備考」を参考にして、屋内、屋外、不明のいずれかを設定してください。
    - 「数量」は「台数」の値を使用してください。
    - メモにはその他の項目にない情報（例えば、冷却能力、加熱能力、風量、消費電力）を記載してください。
###設備種類、セクション、サブセクションの決め方
1.まず機器情報をもとに以下のキーワードから今回最適な項目を必ず選んでください
キーワード：
    • 冷温水発生機
    • 冷凍機
    • 外調機
    • 排気ファン
    • 給気ファン
    • 全熱交換器
    • FCU
2.選んだキーワードに対応する「設備種類」、「セクション」、「サブセクション」を値として入力してください。
- キーワード：冷温水発生機
    - 設備種類：空調設備
    - セクション：熱源機器
    - サブセクション：冷温水発生器
- キーワード：冷凍機
    - 設備種類：空調設備
    - セクション：熱源機器
    - サブセクション：冷凍機
- キーワード：外調機
    - 設備種類：空調設備
    - セクション：空調機
    - サブセクション：外調機
- キーワード：排気ファン
    - 設備種類：空調設備
    - セクション：：給排気設備
    - サブセクション：排気ファン
- キーワード：給気ファン
    - 設備種類：空調設備
    - セクション：給排気設備
    - サブセクション：給気ファン
- キーワード：全熱交換器
    - 設備種類：空調設備
    - セクション：給排気設備
    - サブセクション：全熱交換器
- キーワード：FCU
    - 設備種類：空調設備
    - セクション：空調機
    - サブセクション：FCU

###jsonデータ
以下のJSONデータを基に処理してください: {input_json}
    """

    chat_completion = openai.ChatCompletion.create(
            messages=[{
                "role": "user",
                "content": prompt
            }],
            model="gpt-4o",
            temperature=0.1,
            response_format={"type": "json_object"}
        )
    a=json.loads(chat_completion.choices[0].message.content)

    wb = Workbook()
    ws = wb.active
    ws.title = "機器情報"

    # 列名を定義
    columns = ["機器名", "メーカー", "設備種類", "セクション", "サブセクション", "型式", "型番", "設置場所", "屋内・屋外", "数量", "メモ"]
    # 列名を1行目に設定
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # データを行ごとに記入
    current_row = 2  # データは2行目から記入
    for machine in a['機器情報']:
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=current_row, column=col_idx, value=machine.get(col_name, ""))
        current_row += 1

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

if __name__ == "__main__":
    st.title("設備データ処理アプリケーション")
    uploaded_file = st.file_uploader("Excelファイルをアップロードしてください", type=['xlsx'])

    if uploaded_file is not None:
        # セッション状態に保存されていない場合のみ処理を実行
        if "sheet1_buffer" not in st.session_state or "sheet2_buffer" not in st.session_state:
            with st.spinner('処理中...'):
                # 一時ファイルを作成してパスを取得
                with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(uploaded_file.name)[1]) as tmp_file:
                    tmp_file.write(uploaded_file.getvalue())
                    temp_path = tmp_file.name

                try:
                    # シート処理
                    st.session_state["sheet1_buffer"] = sheet1(temp_path)
                    st.session_state["sheet2_buffer"] = sheet2(temp_path)
                finally:
                    # 一時ファイルのクリーンアップ
                    if os.path.exists(temp_path):
                        os.unlink(temp_path)

        # ダウンロードボタン
        st.write("以下のExcelファイルへ転記が完了しました。ダウンロードしてください。")
        st.download_button(
            "シート1", 
            st.session_state["sheet1_buffer"], 
            file_name="sheet1.xlsx", 
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.download_button(
            "シート2", 
            st.session_state["sheet2_buffer"], 
            file_name="sheet2.xlsx", 
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
